# -*- coding: utf-8 -*-
"""
Created on Thu Jan 20 13:30:46 2022

@author: kshaad
"""

import math
import streamlit as st
from openpyxl import load_workbook

nGauge=0
para_list=[]
para_per_gauge=[]

methdStr=''
blnReady2calc= False
blnInputLoad = False
class para:
    case=0
    ta=0.
    tb=0.   
    def __init__(self, nm,gauge,unit,ts,val):
        self.nm=nm
        self.gaugeID=gauge
        self.unit=unit
        self.ts=ts
        self.val=val
        self.ex=val
        
st.title("FHI: Water Quality Index")
with st.expander("Input requirements", expanded=False):
    st.markdown("""
                **Provide input data in following format.** [See examples here](https://github.com/freshwaterhealth/fhiScripts/blob/main/01SampleDatasets/02_Ecosystem%20Vitality/Water%20Quality%20Mock%20Data.xlsx)
                    
                    1. Input is accepted in excel spreadsheet (XLSX) format.
                    2. Thresholds can be entered via the interface below 
        """)
st.sidebar.title("Inputs")
st.sidebar.subheader("Step 1: Upload Water quality data")
uploaded_file = st.sidebar.file_uploader("Upload XLSX file",type=['xlsx'])
if uploaded_file is not None:
    try:
        wb = load_workbook(uploaded_file)
        gauge_list=wb.sheetnames
        nGauge=len(wb.sheetnames)
        # wb.active=2   --> Method to change sheet names    

        for ig in range(0,nGauge):
            wb.active=ig
            ws1 = wb.active    
            nRow=ws1.max_row
            nCol=ws1.max_column    
            nPara_thisGauge=nCol-1
            
            for i in range(2,nCol+1):
                t=[]
                v=[]
                n=ws1.cell(row=1,column=i).value
                u=ws1.cell(row=2,column=i).value
                for j in range(3,nRow+1):
                    tt=ws1.cell(row=j,column=1).value
                    vt=ws1.cell(row=j,column=i).value
                    if vt!=None:
                        v.append(float(vt))
                        t.append(tt)
                para_list.append(para(n,gauge_list[ig],u,t,v))
                #st.write(n)
                #st.write(gauge_list[ig])
                #st.write(t)
                #st.write(v)
            
            blnInputLoad = True          
    except:
            st.error("Faliure loading water quality data")
            blnInputLoad = False
#Select method and display some stats on input data
methdStr = st.select_slider(
     'Select the method for Ecosystem Service Indicator calculation',
     options=['Method 1', 'Method 2', 'Method 3'], value='Method 2')

uploaded_file = st.sidebar.file_uploader("Upload threshold data",type=['xlsx']) 
st.sidebar.markdown("""[See template here](https://github.com/freshwaterhealth/streamlit-soc/blob/main/templates/wqThreholdsTemplate.xlsx)""")
try:
    if uploaded_file is not None:
        wb = load_workbook(uploaded_file)
        if(len(wb.sheetnames)>1):
            st.error("Expecting one sheet, found multiple")
            st.stop()
            
        ws1 = wb.active 

        nRow=ws1.max_row
        nCol=ws1.max_column   

        for i in range(2,nRow+1):

            p=ws1.cell(row=i,column=1).value
            syb=ws1.cell(row=i,column=2).value
            v1=0.
            v2=0.
            if syb=='<':
                case=1
                v1=ws1.cell(row=i,column=3).value
            elif syb=='>':
                case=2
                v1=ws1.cell(row=i,column=3).value
            elif syb=="<>":
                case=3
                v1=ws1.cell(row=i,column=3).value
                v2=ws1.cell(row=i,column=4).value
            else:
                st.error("Expecting one of '<','>' or '<>'")
                st.stop()
            for obj in para_list:
                if obj.nm==p:
                    obj.case=case
                    obj.ta=float(v1)
                    obj.tb=float(v2)        
        blnReady2calc = True
except:
    st.error("Faliure loading thresholds")
    blnReady2calc = False

if st.button('Calculate WQT'):
    def retExcursion(v,c,ta,tb):
        if c==1:
            if v<ta:
                return ((ta/v)-1)
            else:
                return (0.)
        if c==2:
            if v>ta:
                return ((v/ta)-1)
            else:
                return (0.)
        if c==3:
            if v<ta:
                return ((ta/v)-1)
            elif v>tb:
                return ((v/tb)-1)
            else:
                return 0.
    
    if blnReady2calc and blnInputLoad:
        currGauge=para_list[0].gaugeID
        WQT=[]
        F1cnt=0.
        F1fail=0.        
        F2cnt=0.
        F2fail=0.
        se=0.
        nse=0.
        mse=0.
        for obj in para_list:
            if obj.gaugeID == currGauge:
                F1cnt+=1
                F2cnt+=len(obj.val)
            else:
                #Calcuate indicator for previous gauge
                F1=100*(F1fail/F1cnt)
                F2=100*(F2fail/F2cnt)
                nse=se/F2cnt
                if (F2cnt>0):
                    mse=se/F2cnt
                else:
                    mse=0
                F31=100*(nse/(nse+1))
                F32=100*(mse/(mse+1))
                if methdStr=="Method 1":
                    wq=100 -math.sqrt((F1*F1+F2*F2+F31*F31)/3)
                elif methdStr=="Method 2":
                    wq=100 - math.sqrt(F1*F31) 
                else:
                    wq=100-math.pow(F1*F2*F32,(1/3))
                WQT.append(wq)
                    
                #Reset everything
                currGauge=obj.gaugeID
                F1cnt=1.
                F1fail=0.       
                F2cnt=len(obj.val)
                F2fail=0.
                se=0.
                nse=0.
                mse=0.
                
            if obj.case==0:
                st.error("At least one parameter with Thresolds:"+obj.gaugeID+","+obj.nm)                
                st.stop()
            for i in range(0,len(obj.val)-1):
                #st.write("Value:"+str(obj.val[i])+","+str(obj.case)+","+str(obj.ta)+","+str(obj.tb))
                obj.ex[i]=retExcursion(obj.val[i],obj.case,obj.ta,obj.tb)
                blnF1=False
                se=se+obj.ex[i]
                if obj.ex[i]>0:
                    F2fail+=1
                    blnF1=True
            if blnF1:
                F1fail=F1fail+1
        #Calcuate for last gauge
        F1=100*(F1fail/F1cnt)
        F2=100*(F2fail/F2cnt)
        nse=se/F2cnt
        if (F2cnt>0):
            mse=se/F2cnt
        else:
            mse=0
        F31=100*(nse/(nse+1))
        F32=100*(mse/(mse+1))
        if methdStr=="Method 1":
            wq=100 -math.sqrt((F1*F1+F2*F2+F31*F31)/3)
        elif methdStr=="Method 2":
            wq=100 - math.sqrt(F1*F31) 
        else:
            wq=100-math.pow(F1*F2*F32,(1/3))
        WQT.append(wq)
        
        st.success("Indicator Score:")
        for ig in range(0,nGauge):
            st.metric(gauge_list[ig], round(WQT[ig],2), "")