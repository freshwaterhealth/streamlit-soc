# -*- coding: utf-8 -*-
"""
Created on Thu Jan 20 13:30:46 2022

@author: kshaad
"""

import math
import streamlit as st
from openpyxl import load_workbook

nSU=0
su_list=[]
level=['F1','F2','F3Fuzzy','F3Sharp']
lvlStr=''
methdStr=''
blnReady2calc= False
blnInputLoad = False
class su:
    case=1
    ta=0.
    tb=0.
    threType=0    
    def __init__(self, nm, unit,ts,val):
        self.nm=nm
        self.unit=unit
        self.ts=ts
        self.val=val
        self.ex=val
        self.thres=[]
        self.exc=0.
        
st.title("FHI: Freshwater Ecosystem Services")
with st.expander("Input requirements", expanded=False):
    st.markdown("""
                **Provide input data in following format.** [See examples here](https://github.com/freshwaterhealth/fhiScripts/tree/main/01SampleDatasets/03_Ecosystem%20Services)
                    
                    1. Input is accepted in excel spreadsheet (XLSX) format.
                    2. If providing thresholds as a excel spreadsheet as well, use the same template with:
                        2.1 In row 2, replace 'units' with conditional: <,> or <>
                        2.2 Replace the supply values with non-compliance threshold value
                        2.3 For outside range (<>) conditional provide value pair is comma seperated. 
        """)
st.sidebar.title("Inputs")
st.sidebar.subheader("Step 1: Set water-related Ecosystem Service type")
es_name = st.sidebar.selectbox(
         'Ecosystem Service type:',
         ('Water supply reliability', 'Biomass for consumption', 'Sediment regulation','Deviation of water quality metrics','Flood regulation','Exposure to water-associated diseases'))
st.subheader("Setup for "+ es_name)

st.sidebar.subheader("Step 2: Upload Ecosystem Service supply data")
uploaded_file = st.sidebar.file_uploader("Upload XLSX file",type=['xlsx'])
if uploaded_file is not None:
    try:
        wb = load_workbook(uploaded_file)
        #st.write(wb.sheetnames)
        # wb.active=2   --> Method to change sheet names
        if(len(wb.sheetnames)>1):
            st.error("Expecting one sheet, found multiple")
            st.stop()
            
        ws1 = wb.active    
        lvlStr=ws1.cell(row=1,column=1).value
        if any(x in level for x in level):
            st.write("Calculation level set as: ",lvlStr)
        else:
            st.write(lvlStr)
            st.error("Calculation level specified does not match standard list")
            st.stop()
        nRow=ws1.max_row
        nCol=ws1.max_column    
        nSU=nCol-1
        st.write("Number of Spatial Units (SU):",nSU)
        for i in range(2,nCol+1):
            t=[]
            v=[]
            n=ws1.cell(row=1,column=i).value
            u=ws1.cell(row=2,column=i).value
            for j in range(3,nRow+1):
                tt=ws1.cell(row=j,column=1).value
                vt=ws1.cell(row=j,column=i).value
                if vt!=None:
                    v.append(float(vt))
                    t.append(tt)
            su_list.append(su(n,u,t,v))
            blnInputLoad = True
            #st.write(t)
            #st.write(v)
    except:
            st.error("Faliure loading thresholds")
            blnInputLoad = False
#Select method and display some stats on input data
methdStr = st.select_slider(
     'Select the method for Ecosystem Service Indicator calculation',
     options=['Method 1', 'Method 2', 'Method 3'], value='Method 2')

#Interactive demand threshold based on choices so far        
if lvlStr=="F1" or lvlStr=="F2" or lvlStr=="F3Fuzzy":
    blnReady2calc = True
elif lvlStr=="F3Sharp":
    st.subheader("Set Demand Thresholds")
    opt=st.radio("Select threshold type: " , ('Apply single threshold to all SU','Apply different thresholds to each SU', 'Upload different thresolds to each SU and time step'))
    if opt == 'Apply single threshold to all SU':
        st.write("**Spatial Unit is non-compliant when:** ")
        col1, col2 = st.columns(2)
        option = col1.selectbox(
         'Supply is',
         ('Less than (<)', 'Greater than (>)', 'Outside range (<>)'))
        if option=='Outside range (<>)':
            su.ta=col2.number_input("Less than")
            su.tb=col2.number_input("Or greater than")
            su.case=3
        else:
            su.ta=col2.number_input("Threshold")
            if option=='Less than (<)':
                su.case=1
            else:
                su.case=2
        su.threType=0
        blnReady2calc = True
    
    if opt == 'Apply different thresholds to each SU':
        for obj in su_list:
            tempStr= "Spatial Unit " +str(obj.nm)+" is non-compliant when:"
            st.write("**"+tempStr+"**")
            col1, col2 = st.columns(2)
            option = col1.selectbox(
             'Supply is',
             ('Less than (<)', 'Greater than (>)', 'Outside range (<>)'),
             key=obj.nm)
            if option=='Outside range (<>)':
                obj.ta=col2.number_input("Less than", key=obj.nm)
                obj.tb=col2.number_input("Or greater than", key=obj.nm)
                obj.case=3
            else:
                obj.ta=col2.number_input("Threshold ("+str(obj.unit)+")", key=obj.nm)
                if option=='Less than (<)':
                    obj.case=1
                else:
                    obj.case=2
        su.threType=0
        blnReady2calc = True
    
    if opt == 'Upload different thresolds to each SU and time step':        
        uploaded_file = st.file_uploader("Upload threshold data",type=['xlsx']) 
        st.markdown("""[See template here](https://www.example.com)""")
        try:
            if uploaded_file is not None:
                wb = load_workbook(uploaded_file)
                if(len(wb.sheetnames)>1):
                    st.error("Expecting one sheet, found multiple")
                    st.stop()
                    
                ws1 = wb.active    
                nRow=ws1.max_row
                nCol=ws1.max_column    
                for i in range(2,nCol+1):
                    syb=ws1.cell(row=2,column=i).value
                    if syb=='<':
                        su_list[i-2].case=1
                    elif syb=='>':
                        su_list[i-2].case=2
                    elif syb=="<>":
                        su_list[i-2].case=3
                    else:
                        st.error("Expection one of '<','>' or '<>'")
                        st.stop()
                    for j in range(3,nRow+1):
                        th=ws1.cell(row=j,column=i).value
                        if th!=None:
                            su_list[i-2].thres.append(ws1.cell(row=j,column=i).value)
                su.threType=1
                blnReady2calc = True
        except:
            st.error("Faliure loading thresholds")
            blnReady2calc = False

#Calculation and results
if st.button('Calculate ESI'):
    nF1=0.
    nF2=0.
    nF2len=0.
    suF3=0.
    F1=0.
    F2=0.
    F31=0.
    F32=0.
    ESIm1=0.
    ESIm2=0.
    ESIm3=0.
    
    def retExcursion(v,c,ta,tb):
        if c==1:
            if v<ta:
                return ((ta/v)-1)
            else:
                return (0.)
        if c==2:
            if v>ta:
                return ((v/ta)-1)
            else:
                return (0.)
        if c==3:
            if v<ta:
                return ((ta/v)-1)
            elif v>tb:
                return ((v/tb)-1)
            else:
                return 0.
            
        
    if blnReady2calc and blnInputLoad:
        if lvlStr=="F1":
            for obj in su_list:
                nF1=nF1+obj.val[0]
            F1=100*((nSU-nF1)/nSU)
            ESIm2=100-F1
        if lvlStr=="F2":
            for obj in su_list:
                suF2len=len(obj.val)
                suF2cnt=0.
                for i in range(0,len(obj.val)-1):
                    suF2cnt=suF2cnt+obj.val[i]
                if suF2cnt<suF2len:
                    nF1=nF1+1
                nF2+=suF2len-suF2cnt
                nF2len+=suF2len
            F1=100*(nF1/nSU)
            F2=100*(nF2/nF2len)
            ESIm2=100 - math.sqrt(F1*F2)
        if lvlStr=="F3Fuzzy":
            for obj in su_list:
                suF2len=len(obj.val)
                suF2cnt=0.
                suF3sum=0.
                for i in range(0,len(obj.val)-1):
                    suF3sum=suF3sum+obj.val[i]
                    if obj.val[i]>0:
                        suF2cnt+=1
                if suF2cnt>0:
                    nF1=nF1+1
                nF2+=suF2cnt
                nF2len+=suF2len
                suF3+=suF3sum
            F1=100*(nF1/nSU)
            F2=100*(nF2/nF2len)
            nse=suF3/nF2len
            F31=100*(nse/(nse+1))
            if nF2>0:
                mse=suF3sum/nF2
            else:
                mse=0.
            F32=100*(mse/(mse+1))
            ESIm2=100 - math.sqrt(F1*F31)
            ESIm3=100-math.pow(F1*F2*F32,(1/3))
                
        if lvlStr=="F3Sharp":
            #first convert obj.val to an excursion ts
            if su.threType==0:
                for obj in su_list:
                    for i in range(0,len(obj.val)-1):
                        obj.ex[i]=retExcursion(obj.val[i],obj.case,obj.ta,obj.tb)
            else:
                for obj in su_list:
                    for i in range(0,len(obj.val)-1):
                        if obj.case==1 or obj.case==2:
                            obj.ex[i]=retExcursion(obj.val[i],obj.case,float(obj.thres[i]),0.)
                        else:
                            tlist = obj.thres[i].split(",")
                            obj.ex[i]=retExcursion(obj.val[i],obj.case,float(tlist[0]),float(tlist[1]))
            #then same as F3Fuzzy
            for obj in su_list:
                suF2len=len(obj.ex)
                suF2cnt=0.
                suF3sum=0.
                for i in range(0,len(obj.ex)-1):
                    suF3sum=suF3sum+obj.ex[i]
                    if obj.ex[i]>0:
                        suF2cnt+=1
                if suF2cnt>0:
                    nF1=nF1+1
                nF2+=suF2cnt
                nF2len+=suF2len
                suF3+=suF3sum
            F1=100*(nF1/nSU)
            F2=100*(nF2/nF2len)
            nse=suF3/nF2len
            F31=100*(nse/(nse+1))
            if nF2>0:
                mse=suF3sum/nF2
            else:
                mse=0.
            F32=100*(mse/(mse+1))
            ESIm1=100 -math.sqrt((F1*F1+F2*F2+F31*F31)/3)
            ESIm2=100 - math.sqrt(F1*F31)
            ESIm3=100-math.pow(F1*F2*F32,(1/3))
        #Display results
        st.success("Indicator Score:")
        if methdStr=="Method 1":
            if lvlStr=="F3Sharp":
                st.metric("ESIm1", round(ESIm1,2), "")
            else:
                st.info("Method 1 only supports F3Sharp level of evidence")
        elif methdStr=="Method 2":
            st.metric("ESIm2", round(ESIm2,2), "")
        else:
            st.metric("ESIm3", round(ESIm3,2), "")
    else:
         st.info("Some inputs still missing!")
    
